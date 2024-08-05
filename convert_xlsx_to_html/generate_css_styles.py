import openpyxl
from openpyxl.utils import get_column_letter

def generate(ws):
    css_styles = generate_table_styles()

    for row in ws.iter_rows():
        for cell in row:
            css_styles += generate_cell_styles(ws, cell)
            css_styles += generate_merged_cell_styles(ws, cell)

    return css_styles

def generate_table_styles():
    return """
.excel-table {
    border-collapse: collapse;
}
"""

def generate_cell_styles(ws, cell):
    css_properties = []

    css_properties.append(generate_bold_styles(cell))
    css_properties.append(generate_border_styles(cell))
    css_properties.append(generate_width_styles(ws, cell))
    css_properties.append(generate_height_styles(ws, cell))
    css_properties.append(generate_font_styles(cell))
    css_properties.append(generate_alignment_styles(cell))
    css_properties = list(filter(lambda x: x != "", css_properties))

    css_indent = "    "
    css_class = f".cell-{cell.row}-{cell.column} {{\n{css_indent}"
    css_class += f"\n{css_indent}".join(css_properties)
    css_class += "\n}\n"
    return css_class

def generate_bold_styles(cell):
    if cell.font and cell.font.bold:
        return "font-weight: bold;"

    return ""

def generate_border_styles(cell):
    border_sides = ["top", "left", "right", "bottom"]
    border_styles = []

    for border_side in border_sides:
        border = getattr(cell.border, border_side)
        if border.style and border.style != "none":
            border_color = "000000"
            if border.color:
                if len(str(border.color.rgb)) == 6:
                    border_color =  str(border_color.rgb) # Default color is black
            border_styles.append(f"border-{border_side}: solid 1px #{border_color};")

    if border_styles:
        return " ".join(border_styles)

    return ""

def generate_width_styles(ws, cell):
    col_width = ws.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width
    col_width = points_to_pixels(col_width)
    return f"width: {col_width}px;"

def generate_height_styles(ws, cell):
    row_height = ws.row_dimensions[cell.row].height
    return f"height: {row_height}px;"

def generate_font_styles(cell):
    return f"font-size: {cell.font.size}px;"

def generate_alignment_styles(cell):
    horizontal_alignment = cell.alignment.horizontal
    if horizontal_alignment in ["left", "center", "right"]:
        return f"text-align: {horizontal_alignment};"

    return ""

def generate_merged_cell_styles(ws, cell):
    css_styles = ""
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate not in merged_range:
            continue

        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if row != merged_range.min_row or col != merged_range.min_col:
                    merged_cell_class = f".cell-{row}-{col}"
                    css_styles += f"{merged_cell_class} {{ display: none; }}\n"
    return css_styles

def points_to_pixels(points):
    pixels = points * 6.665  #1.33
    return round(pixels)
